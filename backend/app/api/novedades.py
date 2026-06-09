import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from app.db.database import get_db

router_yaguar = APIRouter(prefix="/yaguar/novedades", tags=["Yaguar - Novedades"])
router_diarco = APIRouter(prefix="/diarco/novedades", tags=["Diarco - Novedades"])

TIPOS_VALIDOS = ("devolucion", "faltante", "cambio")


class NovedadIn(BaseModel):
    semana: str
    cod_bar: Optional[str] = None
    cod_art: Optional[str] = None
    descrip: Optional[str] = None
    cliente: Optional[str] = None
    cliente_nombre: Optional[str] = None
    tipo: str
    observaciones: Optional[str] = None
    unidades: int = 0
    bultos: int = 0
    uxb: int = 0
    precio: Optional[float] = None


# ── Helpers ────────────────────────────────────────────────────────────────────

def _list(mayorista: str, semana: str):
    with get_db() as cur:
        cur.execute("""
            SELECT id, cod_bar, cod_art, descrip, cliente, cliente_nombre,
                   tipo, observaciones, unidades, bultos, uxb, created_at
            FROM novedades
            WHERE mayorista = %s AND semana = %s
            ORDER BY created_at DESC
        """, (mayorista, semana))
        return [dict(r) for r in cur.fetchall()]


def _add(mayorista: str, data: NovedadIn):
    if data.tipo not in TIPOS_VALIDOS:
        raise HTTPException(400, f"Tipo inválido. Valores: {', '.join(TIPOS_VALIDOS)}")
    if not data.cod_art and not data.cod_bar:
        raise HTTPException(400, "Se requiere cod_art o cod_bar")
    if not data.cliente:
        raise HTTPException(400, "Se requiere un cliente")
    if not data.semana:
        raise HTTPException(400, "Se requiere una semana")
    with get_db() as cur:
        cur.execute("""
            INSERT INTO novedades
                (mayorista, semana, cod_bar, cod_art, descrip, cliente, cliente_nombre,
                 tipo, observaciones, unidades, bultos, uxb, precio)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id, cod_bar, cod_art, descrip, cliente, cliente_nombre,
                      tipo, observaciones, unidades, bultos, uxb, precio, created_at
        """, (
            mayorista, data.semana,
            (data.cod_bar or "").strip() or None,
            (data.cod_art or "").strip() or None,
            (data.descrip or "").strip() or None,
            data.cliente,
            data.cliente_nombre,
            data.tipo,
            (data.observaciones or "").strip() or None,
            max(0, data.unidades),
            max(0, data.bultos),
            max(0, data.uxb),
            data.precio if data.precio and data.precio > 0 else None,
        ))
        return dict(cur.fetchone())


def _delete(mayorista: str, item_id: int):
    with get_db() as cur:
        cur.execute(
            "DELETE FROM novedades WHERE id = %s AND mayorista = %s RETURNING id",
            (item_id, mayorista)
        )
        if not cur.fetchone():
            raise HTTPException(404, "Novedad no encontrada")
    return {"ok": True}


def _lookup(mayorista: str, cod_bar: str, semana: Optional[str]):
    with get_db() as cur:
        for col in ("cod_bar", "cod_bar_bulto"):
            q = f"SELECT cod_art, descrip, uxb FROM pick WHERE {col}=%s AND mayorista=%s"
            params = [cod_bar, mayorista]
            if semana:
                q += " AND semana=%s"
                params.append(semana)
            q += " LIMIT 1"
            cur.execute(q, params)
            row = cur.fetchone()
            if row:
                return {"cod_art": row["cod_art"], "descrip": row["descrip"],
                        "uxb": row["uxb"] or 0, "found": True}
    return {"cod_art": None, "descrip": None, "uxb": 0, "found": False}


def _search(mayorista: str, q: str, semana: Optional[str]):
    with get_db() as cur:
        pattern = f"%{q}%"
        query = """
            SELECT DISTINCT ON (cod_art) cod_bar, cod_art, descrip, uxb
            FROM pick WHERE (descrip ILIKE %s OR cod_art ILIKE %s) AND mayorista=%s
        """
        params = [pattern, pattern, mayorista]
        if semana:
            query += " AND semana=%s"
            params.append(semana)
        query += " ORDER BY cod_art, descrip LIMIT 20"
        cur.execute(query, params)
        return [dict(r) for r in cur.fetchall()]


def _export(mayorista: str, semana: str):
    with get_db() as cur:
        cur.execute("""
            SELECT n.semana, n.tipo, n.cliente_nombre, n.cliente, n.cod_art, n.descrip,
                   n.bultos, COALESCE(n.uxb, ap.uxb, 0) AS uxb, n.unidades, n.observaciones,
                   COALESCE(n.precio, ap.precio_con_iva) AS precio_unit
            FROM novedades n
            LEFT JOIN articulos_precios ap ON ap.cod_art = n.cod_art AND ap.mayorista = n.mayorista
            WHERE n.mayorista = %s AND n.semana = %s
            ORDER BY n.created_at DESC
        """, (mayorista, semana))
        rows = [dict(r) for r in cur.fetchall()]

    from app.api.excel_theme import (
        hdr_cell, data_cell, make_table, set_col_widths, build_filename, stream_wb, MONEY, ROW_ALT, WHITE
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Novedades"

    hdrs = [
        ("Cód. Cliente",     14),
        ("Cliente",          28),
        ("Cód. Artículo",   14),
        ("Descripción",      40),
        ("Unidades totales", 14),
        ("Tipo",             14),
        ("Observaciones",    30),
    ]
    for ci, (label, width) in enumerate(hdrs, 1):
        ws.cell(row=1, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28

    tipo_labels = {"devolucion": "Devolución", "faltante": "Faltante", "cambio": "Cambio"}
    for ri, r in enumerate(rows, 2):
        uxb       = r["uxb"] or 0
        uni_total = (r["bultos"] or 0) * uxb + (r["unidades"] or 0)
        tipo_str  = tipo_labels.get(r["tipo"], r["tipo"])

        _f = openpyxl.styles.PatternFill("solid", fgColor=ROW_ALT if ri % 2 == 0 else WHITE)
        for ci in range(1, 8):
            ws.cell(row=ri, column=ci).fill = _f
        data_cell(ws, ri, 1, r["cliente"],             align="center")
        data_cell(ws, ri, 2, r["cliente_nombre"],      align="left")
        data_cell(ws, ri, 3, r["cod_art"],             align="center")
        data_cell(ws, ri, 4, r["descrip"],             align="left")
        data_cell(ws, ri, 5, uni_total,                align="center")
        data_cell(ws, ri, 6, tipo_str,                 align="center", bold=True)
        data_cell(ws, ri, 7, r["observaciones"] or "", align="left")

    ws.auto_filter.ref = f"A1:G{len(rows)+1}" if rows else None
    ws.freeze_panes = "A2"
    fname = build_filename("Novedades", f"{semana} {mayorista.capitalize()}")
    return stream_wb(wb, fname)


# ── Rutas Yaguar ───────────────────────────────────────────────────────────────

@router_yaguar.get("/")
def yaguar_list(semana: str = Query(...)):
    return _list("yaguar", semana)

@router_yaguar.post("/")
def yaguar_add(data: NovedadIn):
    return _add("yaguar", data)

@router_yaguar.delete("/{item_id}")
def yaguar_delete(item_id: int):
    return _delete("yaguar", item_id)

@router_yaguar.get("/lookup/{cod_bar}")
def yaguar_lookup(cod_bar: str, semana: Optional[str] = Query(None)):
    return _lookup("yaguar", cod_bar, semana)

@router_yaguar.get("/search")
def yaguar_search(q: str = Query(..., min_length=2), semana: Optional[str] = Query(None)):
    return _search("yaguar", q, semana)

@router_yaguar.get("/export")
def yaguar_export(semana: str = Query(...)):
    return _export("yaguar", semana)


# ── Rutas Diarco ───────────────────────────────────────────────────────────────

@router_diarco.get("/")
def diarco_list(semana: str = Query(...)):
    return _list("diarco", semana)

@router_diarco.post("/")
def diarco_add(data: NovedadIn):
    return _add("diarco", data)

@router_diarco.delete("/{item_id}")
def diarco_delete(item_id: int):
    return _delete("diarco", item_id)

@router_diarco.get("/lookup/{cod_bar}")
def diarco_lookup(cod_bar: str, semana: Optional[str] = Query(None)):
    return _lookup("diarco", cod_bar, semana)

@router_diarco.get("/search")
def diarco_search(q: str = Query(..., min_length=2), semana: Optional[str] = Query(None)):
    return _search("diarco", q, semana)

@router_diarco.get("/export")
def diarco_export(semana: str = Query(...)):
    return _export("diarco", semana)
