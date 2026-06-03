import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from app.db.database import get_db

router_yaguar = APIRouter(prefix="/yaguar/sobrantes", tags=["Yaguar - Sobrantes"])
router_diarco = APIRouter(prefix="/diarco/sobrantes", tags=["Diarco - Sobrantes"])
router_shared = APIRouter(prefix="/sobrantes", tags=["Sobrantes - Compartido"])


class ItemIn(BaseModel):
    cod_bar: Optional[str] = None
    cod_art: Optional[str] = None
    descrip: Optional[str] = None
    mayorista: Optional[str] = None
    precio_unit: Optional[float] = None
    uxb: Optional[int] = None

class CantidadIn(BaseModel):
    unidades: int = 0
    bultos: int = 0

class ListaIn(BaseModel):
    nombre: str


# ── helpers ───────────────────────────────────────────────────────────────────

def _list_listas(mayorista: str):
    with get_db() as cur:
        cur.execute("""
            SELECT lista, COUNT(*) AS items, MAX(created_at) AS ultima
            FROM sobrantes WHERE mayorista = %s
            GROUP BY lista ORDER BY ultima DESC
        """, (mayorista,))
        return [dict(r) for r in cur.fetchall()]


def _get_items(lista: str, mayorista: str):
    with get_db() as cur:
        cur.execute("""
            SELECT id, cod_bar, cod_art, descrip, unidades, bultos
            FROM sobrantes WHERE lista = %s AND mayorista = %s
            ORDER BY created_at DESC
        """, (lista, mayorista))
        return [dict(r) for r in cur.fetchall()]


def _add_item(lista: str, data: ItemIn, mayorista: str):
    cod_bar = (data.cod_bar or "").strip() or None
    cod_art = (data.cod_art or "").strip() or None
    descrip = (data.descrip or "").strip() or None
    if not cod_bar and not cod_art:
        raise HTTPException(400, "Se requiere cod_bar o cod_art")
    with get_db() as cur:
        # Si ya existe en esta lista, devolver el existente (frontend lo resalta)
        if cod_bar:
            cur.execute(
                "SELECT id, cod_bar, cod_art, descrip, unidades, bultos FROM sobrantes WHERE lista=%s AND mayorista=%s AND cod_bar=%s",
                (lista, mayorista, cod_bar)
            )
        else:
            cur.execute(
                "SELECT id, cod_bar, cod_art, descrip, unidades, bultos FROM sobrantes WHERE lista=%s AND mayorista=%s AND cod_art=%s",
                (lista, mayorista, cod_art)
            )
        existing = cur.fetchone()
        if existing:
            return {"action": "existing", **dict(existing)}
        cur.execute(
            "INSERT INTO sobrantes (cod_bar, cod_art, descrip, unidades, bultos, lista, mayorista, precio_unit, uxb) VALUES (%s,%s,%s,0,0,%s,%s,%s,%s) RETURNING id, cod_bar, cod_art, descrip, unidades, bultos",
            (cod_bar, cod_art, descrip, lista, mayorista, data.precio_unit, data.uxb or 0)
        )
        return {"action": "added", **dict(cur.fetchone())}


def _update_item(lista: str, item_id: int, data: CantidadIn, mayorista: str):
    with get_db() as cur:
        cur.execute(
            "UPDATE sobrantes SET unidades=%s, bultos=%s WHERE id=%s AND lista=%s AND mayorista=%s RETURNING id, cod_bar, cod_art, descrip, unidades, bultos",
            (max(0, data.unidades), max(0, data.bultos), item_id, lista, mayorista)
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Item no encontrado")
        return dict(row)


def _delete_item(lista: str, item_id: int, mayorista: str):
    with get_db() as cur:
        cur.execute("DELETE FROM sobrantes WHERE id=%s AND lista=%s AND mayorista=%s", (item_id, lista, mayorista))
        if cur.rowcount == 0:
            raise HTTPException(404, "Item no encontrado")
    return {"ok": True}


def _delete_lista(lista: str, mayorista: str):
    with get_db() as cur:
        cur.execute("DELETE FROM sobrantes WHERE lista=%s AND mayorista=%s", (lista, mayorista))
    return {"ok": True}


def _lookup(cod_bar: str, mayorista: str):
    with get_db() as cur:
        for col in ("cod_bar", "cod_bar_bulto"):
            cur.execute(f"""
                SELECT p.cod_art, p.descrip, COALESCE(p.uxb, ap.uxb, 0) AS uxb,
                       COALESCE(p.precio_unit, ap.precio_con_iva) AS precio_unit
                FROM pick p
                LEFT JOIN articulos_precios ap ON ap.cod_art = p.cod_art AND ap.mayorista = p.mayorista
                WHERE p.{col}=%s AND p.mayorista=%s LIMIT 1
            """, (cod_bar, mayorista))
            row = cur.fetchone()
            if row:
                return {"cod_art": row["cod_art"], "descrip": row["descrip"],
                        "uxb": row["uxb"] or 0, "precio_unit": row["precio_unit"], "found": True}
    return {"cod_art": None, "descrip": None, "uxb": 0, "precio_unit": None, "found": False}


def _search_descrip(q: str, mayorista: str):
    with get_db() as cur:
        cur.execute(
            "SELECT DISTINCT ON (cod_art) cod_bar, cod_art, descrip, uxb, precio_unit FROM pick WHERE descrip ILIKE %s AND mayorista=%s ORDER BY cod_art, descrip LIMIT 20",
            (f"%{q}%", mayorista)
        )
        return [dict(r) for r in cur.fetchall()]


def _export(lista: str, mayorista: str):
    from app.api.excel_theme import (
        hdr_cell, data_cell, make_table, set_col_widths, build_filename, stream_wb, ROW_ALT, WHITE
    )
    with get_db() as cur:
        cur.execute(
            "SELECT cod_bar, cod_art, descrip, unidades, bultos FROM sobrantes WHERE lista=%s AND mayorista=%s ORDER BY created_at DESC",
            (lista, mayorista)
        )
        rows = [dict(r) for r in cur.fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sobrantes"

    hdrs = [("Cód. de Barra",18),("Cód. Artículo",14),("Descripción",42),("Unidades",12),("Bultos",12)]
    for ci, (label, width) in enumerate(hdrs, 1):
        ws.cell(row=1, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28

    _ov = openpyxl.styles.PatternFill("solid", fgColor=WHITE)
    _ev = openpyxl.styles.PatternFill("solid", fgColor=ROW_ALT)
    for ri, r in enumerate(rows, 2):
        _f = _ev if ri % 2 == 0 else _ov
        for ci, (val, al) in enumerate([
            (r["cod_bar"],"center"),(r["cod_art"],"center"),
            (r["descrip"],"left"),(r["unidades"],"center"),(r["bultos"],"center"),
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = _f
            cell.alignment = openpyxl.styles.Alignment(horizontal=al, vertical="center")

    ws.auto_filter.ref = f"A1:E{len(rows)+1}" if rows else None
    ws.freeze_panes = "A2"
    fname = build_filename("Sobrantes", lista)
    return stream_wb(wb, fname)


# ── Yaguar routes ──────────────────────────────────────────────────────────────

@router_yaguar.get("/listas")
def yaguar_list_listas(): return _list_listas("yaguar")

@router_yaguar.post("/listas")
def yaguar_crear_lista(data: ListaIn):
    nombre = data.nombre.strip()
    if not nombre: raise HTTPException(400, "Nombre vacío")
    return {"lista": nombre}

@router_yaguar.delete("/listas/{lista}")
def yaguar_delete_lista(lista: str): return _delete_lista(lista, "yaguar")

@router_yaguar.get("/lookup/{cod_bar}")
def yaguar_lookup(cod_bar: str): return _lookup(cod_bar, "yaguar")

@router_yaguar.get("/search")
def yaguar_search(q: str): return _search_descrip(q, "yaguar")

@router_yaguar.get("/{lista}/export")
def yaguar_export(lista: str): return _export(lista, "yaguar")

@router_yaguar.get("/{lista}")
def yaguar_get_items(lista: str): return _get_items(lista, "yaguar")

@router_yaguar.post("/{lista}/item")
def yaguar_add_item(lista: str, data: ItemIn): return _add_item(lista, data, "yaguar")

@router_yaguar.put("/{lista}/item/{item_id}")
def yaguar_update_item(lista: str, item_id: int, data: CantidadIn): return _update_item(lista, item_id, data, "yaguar")

@router_yaguar.delete("/{lista}/item/{item_id}")
def yaguar_delete_item(lista: str, item_id: int): return _delete_item(lista, item_id, "yaguar")


# ── Diarco routes ──────────────────────────────────────────────────────────────

@router_diarco.get("/listas")
def diarco_list_listas(): return _list_listas("diarco")

@router_diarco.post("/listas")
def diarco_crear_lista(data: ListaIn):
    nombre = data.nombre.strip()
    if not nombre: raise HTTPException(400, "Nombre vacío")
    return {"lista": nombre}

@router_diarco.delete("/listas/{lista}")
def diarco_delete_lista(lista: str): return _delete_lista(lista, "diarco")

@router_diarco.get("/lookup/{cod_bar}")
def diarco_lookup(cod_bar: str): return _lookup(cod_bar, "diarco")

@router_diarco.get("/search")
def diarco_search(q: str): return _search_descrip(q, "diarco")

@router_diarco.get("/{lista}/export")
def diarco_export(lista: str): return _export(lista, "diarco")

@router_diarco.get("/{lista}")
def diarco_get_items(lista: str): return _get_items(lista, "diarco")

@router_diarco.post("/{lista}/item")
def diarco_add_item(lista: str, data: ItemIn): return _add_item(lista, data, "diarco")

@router_diarco.put("/{lista}/item/{item_id}")
def diarco_update_item(lista: str, item_id: int, data: CantidadIn): return _update_item(lista, item_id, data, "diarco")

@router_diarco.delete("/{lista}/item/{item_id}")
def diarco_delete_item(lista: str, item_id: int): return _delete_item(lista, item_id, "diarco")


# ── Shared helpers (sin filtro por mayorista) ──────────────────────────────────

def _list_listas_shared():
    with get_db() as cur:
        cur.execute("""
            SELECT COALESCE(sl.nombre, s.lista) AS lista,
                   COALESCE(s.items, 0) AS items,
                   COALESCE(s.ultima, sl.created_at) AS ultima
            FROM sobrantes_listas sl
            FULL OUTER JOIN (
                SELECT lista, COUNT(*) AS items, MAX(created_at) AS ultima
                FROM sobrantes GROUP BY lista
            ) s ON s.lista = sl.nombre
            ORDER BY ultima DESC
        """)
        return [dict(r) for r in cur.fetchall()]


def _get_items_shared(lista: str, mayorista: Optional[str] = None):
    with get_db() as cur:
        if mayorista:
            cur.execute("""
                SELECT id, cod_bar, cod_art, descrip, unidades, bultos, mayorista
                FROM sobrantes WHERE lista = %s AND mayorista = %s ORDER BY created_at DESC
            """, (lista, mayorista))
        else:
            cur.execute("""
                SELECT id, cod_bar, cod_art, descrip, unidades, bultos, mayorista
                FROM sobrantes WHERE lista = %s ORDER BY created_at DESC
            """, (lista,))
        return [dict(r) for r in cur.fetchall()]


def _lookup_shared(cod_bar: str):
    with get_db() as cur:
        for m in ("yaguar", "diarco"):
            for col in ("cod_bar", "cod_bar_bulto"):
                cur.execute(f"""
                    SELECT p.cod_art, p.descrip, COALESCE(p.uxb, ap.uxb, 0) AS uxb,
                           COALESCE(p.precio_unit, ap.precio_con_iva) AS precio_unit
                    FROM pick p
                    LEFT JOIN articulos_precios ap ON ap.cod_art = p.cod_art AND ap.mayorista = p.mayorista
                    WHERE p.{col}=%s AND p.mayorista=%s LIMIT 1
                """, (cod_bar, m))
                row = cur.fetchone()
                if row:
                    return {"cod_art": row["cod_art"], "descrip": row["descrip"], "mayorista": m,
                            "uxb": row["uxb"] or 0, "precio_unit": row["precio_unit"], "found": True}
    return {"cod_art": None, "descrip": None, "mayorista": None, "uxb": 0, "precio_unit": None, "found": False}


def _search_shared(q: str, mayorista: Optional[str] = None):
    with get_db() as cur:
        if mayorista:
            cur.execute("""
                SELECT DISTINCT ON (cod_art) cod_bar, cod_art, descrip, mayorista, uxb, precio_unit
                FROM pick WHERE descrip ILIKE %s AND mayorista = %s
                ORDER BY cod_art, descrip LIMIT 20
            """, (f"%{q}%", mayorista))
        else:
            cur.execute("""
                SELECT DISTINCT ON (cod_art) cod_bar, cod_art, descrip, mayorista, uxb, precio_unit
                FROM pick WHERE descrip ILIKE %s
                ORDER BY cod_art, descrip LIMIT 20
            """, (f"%{q}%",))
        return [dict(r) for r in cur.fetchall()]


def _add_item_shared(lista: str, data: ItemIn):
    cod_bar = (data.cod_bar or "").strip() or None
    cod_art = (data.cod_art or "").strip() or None
    descrip = (data.descrip or "").strip() or None
    mayorista = (data.mayorista or "yaguar").strip()
    if not cod_bar and not cod_art:
        raise HTTPException(400, "Se requiere cod_bar o cod_art")
    with get_db() as cur:
        if cod_bar:
            cur.execute(
                "SELECT id, cod_bar, cod_art, descrip, unidades, bultos, mayorista FROM sobrantes WHERE lista=%s AND mayorista=%s AND cod_bar=%s",
                (lista, mayorista, cod_bar)
            )
        else:
            cur.execute(
                "SELECT id, cod_bar, cod_art, descrip, unidades, bultos, mayorista FROM sobrantes WHERE lista=%s AND mayorista=%s AND cod_art=%s",
                (lista, mayorista, cod_art)
            )
        existing = cur.fetchone()
        if existing:
            return {"action": "existing", **dict(existing)}
        cur.execute(
            """INSERT INTO sobrantes (cod_bar, cod_art, descrip, unidades, bultos, lista, mayorista, precio_unit, uxb)
               VALUES (%s,%s,%s,0,0,%s,%s,%s,%s)
               RETURNING id, cod_bar, cod_art, descrip, unidades, bultos, mayorista""",
            (cod_bar, cod_art, descrip, lista, mayorista, data.precio_unit, data.uxb or 0)
        )
        return {"action": "added", **dict(cur.fetchone())}


def _update_item_shared(lista: str, item_id: int, data: CantidadIn):
    with get_db() as cur:
        cur.execute(
            "UPDATE sobrantes SET unidades=%s, bultos=%s WHERE id=%s AND lista=%s RETURNING id, cod_bar, cod_art, descrip, unidades, bultos, mayorista",
            (max(0, data.unidades), max(0, data.bultos), item_id, lista)
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Item no encontrado")
        return dict(row)


def _export_shared(lista: str):
    with get_db() as cur:
        cur.execute("""
            SELECT s.mayorista, s.cod_bar, s.cod_art, s.descrip, s.unidades, s.bultos,
                   COALESCE(s.uxb, ap.uxb, p.uxb, 0) AS uxb,
                   COALESCE(s.precio_unit, ap.precio_con_iva, p.precio_unit) AS precio_unit
            FROM sobrantes s
            LEFT JOIN articulos_precios ap
                ON ap.cod_art = s.cod_art AND ap.mayorista = s.mayorista
            LEFT JOIN LATERAL (
                SELECT uxb, precio_unit FROM pick
                WHERE pick.cod_art = s.cod_art AND pick.mayorista = s.mayorista
                ORDER BY id DESC LIMIT 1
            ) p ON true
            WHERE s.lista = %s
            ORDER BY s.mayorista, s.created_at DESC
        """, (lista,))
        rows = [dict(r) for r in cur.fetchall()]

    from app.api.excel_theme import (
        hdr_cell, data_cell, make_table, set_col_widths, build_filename, stream_wb, MONEY, ROW_ALT, WHITE
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sobrantes"

    hdrs = [
        ("Mayorista",          12), ("Cód. Artículo", 14), ("Descripción",       44),
        ("Unid.",               8), ("Bultos",          8), ("UxB",                7),
        ("Unid. Totales",      14), ("Precio unit. c/IVA", 18), ("Total c/IVA",   16),
    ]
    for ci, (label, width) in enumerate(hdrs, 1):
        ws.cell(row=1, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28

    for ri, r in enumerate(rows, 2):
        precio      = round(float(r["precio_unit"]), 2) if r["precio_unit"] is not None else ""
        uxb         = r["uxb"] or 0
        unid_totales = r["unidades"] + (uxb * r["bultos"]) if uxb else r["unidades"]
        total        = round(float(r["precio_unit"]) * unid_totales, 2) if r["precio_unit"] is not None else ""

        data_cell(ws, ri, 1, r["mayorista"].upper() if r["mayorista"] else "", align="center")
        data_cell(ws, ri, 2, r["cod_art"],    align="center")
        data_cell(ws, ri, 3, r["descrip"],    align="left")
        _f = openpyxl.styles.PatternFill("solid", fgColor=ROW_ALT if ri % 2 == 0 else WHITE)
        for ci, (val, al, fmt) in enumerate([
            (r["mayorista"].upper() if r["mayorista"] else "", "center", None),
            (r["cod_art"],    "center", None),
            (r["descrip"],    "left",   None),
            (r["unidades"],   "center", None),
            (r["bultos"],     "center", None),
            (uxb or "",       "center", None),
            (unid_totales,    "center", None),
            (precio, "right", MONEY if isinstance(precio, float) else None),
            (total,  "right", MONEY if isinstance(total, float) else None),
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = _f
            cell.alignment = openpyxl.styles.Alignment(horizontal=al, vertical="center")
            if fmt: cell.number_format = fmt

    ws.auto_filter.ref = f"A1:I{len(rows)+1}" if rows else None
    ws.freeze_panes = "A2"
    fname = build_filename("Sobrantes", lista)
    return stream_wb(wb, fname)


# ── Shared routes ──────────────────────────────────────────────────────────────

@router_shared.get("/listas")
def shared_list_listas(): return _list_listas_shared()

@router_shared.post("/listas")
def shared_crear_lista(data: ListaIn):
    nombre = data.nombre.strip()
    if not nombre: raise HTTPException(400, "Nombre vacío")
    with get_db() as cur:
        cur.execute(
            "INSERT INTO sobrantes_listas (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING",
            (nombre,)
        )
    return {"lista": nombre}

@router_shared.delete("/listas/{lista}")
def shared_delete_lista(lista: str):
    with get_db() as cur:
        cur.execute("DELETE FROM sobrantes WHERE lista=%s", (lista,))
        cur.execute("DELETE FROM sobrantes_listas WHERE nombre=%s", (lista,))
    return {"ok": True}

@router_shared.get("/lookup/{cod_bar}")
def shared_lookup(cod_bar: str): return _lookup_shared(cod_bar)

@router_shared.get("/search")
def shared_search(q: str, mayorista: Optional[str] = None): return _search_shared(q, mayorista)

@router_shared.get("/{lista}/export")
def shared_export(lista: str): return _export_shared(lista)

@router_shared.get("/{lista}")
def shared_get_items(lista: str, mayorista: Optional[str] = None): return _get_items_shared(lista, mayorista)

@router_shared.post("/{lista}/item")
def shared_add_item(lista: str, data: ItemIn): return _add_item_shared(lista, data)

@router_shared.put("/{lista}/item/{item_id}")
def shared_update_item(lista: str, item_id: int, data: CantidadIn): return _update_item_shared(lista, item_id, data)

@router_shared.delete("/{lista}/item/{item_id}")
def shared_delete_item(lista: str, item_id: int):
    with get_db() as cur:
        cur.execute("DELETE FROM sobrantes WHERE id=%s AND lista=%s", (item_id, lista))
        if cur.rowcount == 0:
            raise HTTPException(404, "Item no encontrado")
    return {"ok": True}
