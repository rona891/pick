import io
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from app.db.database import get_db
from app.api.excel_theme import (
    NAVY, WHITE, ROW_ALT, GOLD, BLUE_MD, BLUE_LT, BORDER, TEXT,
    S_OK, S_PT, S_NO,
    MONEY, PCT, PCT_INT,
    hdr_cell, data_cell, tot_cell, sec_hdr,
    make_table, set_col_widths, build_filename, stream_wb, thin_brd, top_brd,
)

router_yaguar = APIRouter(prefix="/yaguar/export", tags=["Yaguar - Export"])
router_diarco = APIRouter(prefix="/diarco/export", tags=["Diarco - Export"])


# ── Pick export ────────────────────────────────────────────────────────────────

PICK_HEADERS = [
    ("Semana",          "semana",            18),
    ("Código barra",    "cod_bar",           16),
    ("Código art.",     "cod_art",           12),
    ("Descripción",     "descrip",           40),
    ("Cliente",         "nombre",            35),
    ("Zona",            "localidad",         18),
    ("Reparto",         None,                14),
    ("Uni pedidas",     "uni",               12),
    ("Bultos",          "bul",               10),
    ("Uni x bulto",     "uxb",               12),
    ("Uni entregadas",  "cantidad_pickeada", 14),
    ("% entregado",     None,                12),
    ("Estado",          "estado",            22),
    ("Importe pedido",  "importe_total",     16),
]


def _export_picks(semana: str, mayorista: str):
    with get_db() as cur:
        cur.execute("""
            SELECT p.*, COALESCE(z.reparto, '') AS reparto_zona
            FROM pick p
            LEFT JOIN zonas z ON UPPER(p.localidad) = z.nombre
            LEFT JOIN repartos r ON z.reparto = r.nombre
            WHERE p.semana = %s AND p.mayorista = %s
            ORDER BY COALESCE(r.orden, 99) ASC, p.localidad ASC,
                     p.nombre ASC, p.descrip ASC
        """, (semana, mayorista))
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = semana[:31]
    ws.row_dimensions[1].height = 28

    # ── Headers con estilo navy ───────────────────────────────────────────────
    for ci, (label, _, width) in enumerate(PICK_HEADERS, 1):
        hdr_cell(ws, 1, ci, label)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── Fills para filas intercaladas (blanco / azul claro) ───────────────────
    FILL_ODD  = PatternFill("solid", fgColor=WHITE)
    FILL_EVEN = PatternFill("solid", fgColor=ROW_ALT)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, 2):
        r = dict(row)
        uni      = r.get("uni") or 0
        pickeada = r.get("cantidad_pickeada") or 0
        pct      = round(pickeada / uni * 100, 1) if uni > 0 else 0
        reparto  = r.get("reparto_zona") or ""
        estado   = (r.get("estado") or "").lower()

        row_fill = FILL_EVEN if ri % 2 == 0 else FILL_ODD

        if "completado" in estado:
            st_color = S_OK
        elif "entregado" in estado:
            st_color = S_PT
        else:
            st_color = S_NO

        for ci, (label, field, _) in enumerate(PICK_HEADERS, 1):
            if field is None:
                val = reparto if label == "Reparto" else pct
            elif field == "estado":
                val = r.get("estado") or ""
            else:
                val = r.get(field)

            color = st_color if label == "Estado" else TEXT
            fmt   = None
            align = "left"

            if label in ("Uni pedidas", "Bultos", "Uni x bulto", "Uni entregadas"):
                align = "center"
            if label == "% entregado":
                fmt = '0.0"%"'; align = "center"
            if label == "Importe pedido":
                fmt = MONEY; align = "right"

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(bold=(label == "Estado"), size=10, color=color)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if fmt:
                cell.number_format = fmt

    # ── Auto-filtro nativo (compatible con todas las versiones y plataformas) ──
    last_col = get_column_letter(len(PICK_HEADERS))
    last_row = len(rows) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    ws.freeze_panes = "A2"

    fname = build_filename("Pick", semana)
    return stream_wb(wb, fname)


# ── Clientes export ────────────────────────────────────────────────────────────

def _export_clientes(mayorista: str):
    with get_db() as cur:
        cur.execute("""
            SELECT cy.id_yaguar, cy.nombre, cy.localidad, cy.vendedor, cy.flete,
                cy.es_factura_a,
                (
                    SELECT p.semana FROM pick p
                    JOIN semanas s ON s.nombre = p.semana AND s.mayorista = cy.mayorista
                    WHERE p.cliente = cy.id_yaguar AND p.mayorista = cy.mayorista
                    ORDER BY s.created_at DESC LIMIT 1
                ) AS ultimo_pick
            FROM clientes_yaguar cy
            WHERE cy.mayorista = %s
              AND cy.nombre IS NOT NULL AND cy.nombre <> ''
              AND (cy.estado = 'ocupado' OR cy.mayorista = 'diarco')
            ORDER BY cy.nombre
        """, (mayorista,))
        rows = [dict(r) for r in cur.fetchall()]

    es_yaguar = mayorista == "yaguar"
    hdrs = [
        ("Código",       "id_yaguar",    14),
        ("Tipo",         "es_factura_a",  8),
        ("Cliente",      "nombre",       40),
        ("Zona",         "localidad",    22),
        ("Vendedor",     "vendedor",     22),
    ]
    if es_yaguar:
        hdrs.append(("Flete %", "flete", 10))
    hdrs.append(("Último pick", "ultimo_pick", 22))

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.row_dimensions[1].height = 28

    for ci, (label, _, width) in enumerate(hdrs, 1):
        ws.cell(row=1, column=ci, value=label)
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, r in enumerate(rows, 2):
        for ci, (label, field, _) in enumerate(hdrs, 1):
            raw = r.get(field)
            if field == "flete" and raw is not None:
                val   = float(raw)
                fmt   = PCT_INT
                align = "center"
            elif field == "id_yaguar":
                val   = raw or ""
                fmt   = None
                align = "center"
            elif field == "es_factura_a":
                val   = "FA" if raw else "CF"
                fmt   = None
                align = "center"
            elif field == "ultimo_pick":
                val   = raw or "—"
                fmt   = None
                align = "center"
            else:
                val   = raw or ""
                fmt   = None
                align = "left"
            color = S_NO if (field == "ultimo_pick" and not raw) else TEXT
            data_cell(ws, ri, ci, val, align=align, fmt=fmt, color=color)

    if rows:
        last = len(rows) + 1
        make_table(ws, f"A1:{get_column_letter(len(hdrs))}{last}", "TablaClientes")

    ws.freeze_panes = "A2"
    fname = build_filename("Clientes", mayorista.capitalize())
    return stream_wb(wb, fname)


# ── Mod Yaguar export ──────────────────────────────────────────────────────────

def _export_mod_yaguar(semana: str):  # noqa: C901
    with get_db() as cur:
        cur.execute("""
            SELECT cy.id_yaguar AS cod_yag, cy.cod_sis, cy.nombre,
                   cy.telefono, cy.localidad,
                   MAX(p.importe_total) AS total_yaguar,
                   cy.flete AS pct_flete, cy.vendedor
            FROM pick p
            JOIN clientes_yaguar cy
                ON cy.id_yaguar = p.cliente AND cy.mayorista = 'yaguar'
            WHERE p.semana = %s AND p.mayorista = 'yaguar'
              AND cy.nombre IS NOT NULL AND cy.nombre <> ''
            GROUP BY cy.id_yaguar, cy.cod_sis, cy.nombre,
                     cy.telefono, cy.localidad, cy.flete, cy.vendedor
            ORDER BY cy.nombre
        """, (semana,))
        rows = [dict(r) for r in cur.fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "MODELO IA"

    # ── Column widths ─────────────────────────────────────────────────────────
    set_col_widths(ws, {
        1:14, 2:12, 3:10, 4:24, 5:14, 6:14,
        7:16, 8:8,  9:14, 10:14, 11:14, 12:10, 13:10, 14:14, 15:18,
        19:12,
    })

    # ── Row layout ────────────────────────────────────────────────────────────
    n     = len(rows)
    EXTRA = 10          # filas vacías extra para agregar clientes manualmente
    H  = 1
    D1 = 2
    DL = D1 + max(n - 1, 0)       # última fila con datos reales
    DL_EXT = DL + EXTRA             # última fila de la tabla (incluye filas vacías)
    TR = DL_EXT + 1
    TS = TR + 2
    TS1 = TS + 1; TS2 = TS + 2; TS3 = TS + 3; TS4 = TS + 4
    TS5 = TS + 5; TS6 = TS + 6; TS7 = TS + 7; TS8 = TS + 8; TS9 = TS + 9
    CS  = TS + 11; CT = CS + 1; CH = CS + 2
    CD1 = CH + 1;  CDL = CD1 + 89

    # ════════════════════════════════════════════════════════════════════
    # TABLA PRINCIPAL — encabezados
    # ════════════════════════════════════════════════════════════════════
    HDR_MAIN = ["FECHA","COD YAG","COD SIS","NOMBRE","TELEFONO","LOCALIDAD",
                "TOTAL YAGUAR","%","COMISION EBD","ALIMENTOS","TOTAL",
                "FT","BCO","SALDO","VENDEDOR"]
    for ci, label in enumerate(HDR_MAIN, 1):
        hdr_cell(ws, H, ci, label)
    hdr_cell(ws, H, 19, "CLIENTE")
    ws.row_dimensions[H].height = 28

    # ── Datos ─────────────────────────────────────────────────────────────────
    for i, r in enumerate(rows):
        row = D1 + i
        vals = [
            (1,  semana,                                    "left",   None),
            (2,  r["cod_yag"],                              "left",   None),
            (3,  r["cod_sis"],                              "left",   None),
            (4,  (r["nombre"]    or "").strip(),            "left",   None),
            (5,  (r["telefono"]  or "").strip(),            "left",   None),
            (6,  (r["localidad"] or "").strip(),            "left",   None),
            (7,  float(r["total_yaguar"] or 0),             "right",  MONEY),
            (8,  float(r["pct_flete"]    or 0),             "center", PCT_INT),
            (9,  f'=IFERROR(G{row}*H{row},"")',             "right",  MONEY),
            (10, None,                                      "right",  MONEY),
            (11, f'=IFERROR(I{row}+G{row}+J{row},"")',     "right",  MONEY),
            (12, None,                                      "right",  MONEY),
            (13, None,                                      "right",  MONEY),
            (14, f"=+K{row}-L{row}-M{row}",                "right",  MONEY),
            (15, (r["vendedor"] or "").strip(),             "left",   None),
        ]
        for col, val, align, fmt in vals:
            data_cell(ws, row, col, val, align=align, fmt=fmt)
        data_cell(ws, row, 19, f"=IF(M{row}>0,B{row},)", align="center")

    # ── 10 filas vacías extra (con fórmulas para cuando el usuario agrega datos)
    for row in range(DL + 1, DL_EXT + 1):
        data_cell(ws, row, 9,  f'=IFERROR(G{row}*H{row},"")',         align="right", fmt=MONEY)
        data_cell(ws, row, 11, f'=IFERROR(I{row}+G{row}+J{row},"")',  align="right", fmt=MONEY)
        data_cell(ws, row, 14, f"=+K{row}-L{row}-M{row}",             align="right", fmt=MONEY)
        data_cell(ws, row, 19, f"=IF(M{row}>0,B{row},)", align="center")
        ws.cell(row=row, column=8).number_format = PCT_INT  # columna %

    # ── Fila TOTAL ────────────────────────────────────────────────────────────
    if n > 0:
        ws.cell(row=TR, column=6, value="TOTAL")
        ws.cell(row=TR, column=6).font = Font(bold=True, size=10, color=TEXT)
        ws.cell(row=TR, column=6).fill = PatternFill("solid", fgColor=GOLD)
        ws.cell(row=TR, column=6).border = top_brd()
        for col, formula in [
            (7, f"=SUM(G{D1}:G{DL_EXT})"), (9,  f"=SUM(I{D1}:I{DL_EXT})"),
            (10, f"=SUM(J{D1}:J{DL_EXT})"), (11, f'=IFERROR(I{TR}+G{TR}+J{TR},"")'),
            (12, f"=SUM(L{D1}:L{DL_EXT})"), (13, f"=SUM(M{D1}:M{DL_EXT})"),
            (14, f"=SUM(N{D1}:N{DL_EXT})"),
        ]:
            tot_cell(ws, TR, col, formula, fmt=MONEY)

    # ── Excel Table (incluye las 10 filas extra vacías, excluye fila TOTAL) ──
    if n > 0:
        ws.auto_filter.ref = f"A{H}:O{DL_EXT}"
        make_table(ws, f"A{H}:O{DL_EXT}", "TablaModClientes")

    # ════════════════════════════════════════════════════════════════════
    # TOTALES + VENDEDORES
    # ════════════════════════════════════════════════════════════════════
    cell = ws.cell(row=TS, column=1, value="TOTALES")
    cell.font = Font(bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)

    def _tot_lbl(row, a_label, b_label=None):
        ws.cell(row=row, column=1, value=a_label).font = Font(bold=True, size=10, color=TEXT)
        if b_label:
            ws.cell(row=row, column=2, value=b_label).font = Font(bold=True, size=10, color=TEXT)

    def _tot_val(row, col, formula, fmt=MONEY):
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = Font(bold=True, size=10, color=TEXT)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right", vertical="center")

    _tot_lbl(TS1, "TOTAL YAGUAR");          _tot_val(TS1, 3, f"=G{TR}")
    _tot_lbl(TS2, "COMPROBANTES");          _tot_val(TS2, 3, f"=D{CT}")
    _tot_lbl(TS3, "DEVOLUCIONES");          _tot_val(TS3, 3, f"=H{CT}")
    _tot_lbl(TS4, "FT A DEPOSITAR");        _tot_val(TS4, 3, f"=L{TR}")
    _tot_lbl(TS5, "NC ")
    _tot_lbl(TS6, None, "SALDO")
    cell = ws.cell(row=TS6, column=3, value=f"=C{TS1}-C{TS2}-C{TS3}-C{TS4}-C{TS5}")
    cell.font = Font(bold=True, size=10, color="CC0000")
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right", vertical="center")

    # VENDEDORES
    cell = ws.cell(row=TS1, column=6, value="VENDEDORES")
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)

    unique_vendors = list(dict.fromkeys([r["vendedor"] for r in rows if r.get("vendedor")]))
    for vi, vendor in enumerate(unique_vendors):
        vrow = TS2 + vi
        ws.cell(row=vrow, column=6, value=vendor).font = Font(bold=True, size=10)
        cell = ws.cell(row=vrow, column=7,
            value=f'=SUMIF($O${D1}:$O${DL_EXT},"{vendor}",$G${D1}:$G${DL_EXT})')
        cell.font = Font(bold=True, size=10)
        cell.number_format = MONEY
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell = ws.cell(row=vrow, column=8, value=f'=G{vrow}/G{TR}')
        cell.font = Font(bold=True, size=10)
        cell.number_format = PCT_INT
        cell.alignment = Alignment(horizontal="right", vertical="center")

    cell = ws.cell(row=TS8, column=6, value="TOTAL")
    cell.font = Font(bold=True, size=10)
    if n > 0:
        cell = ws.cell(row=TS8, column=7, value=f"=G{TR}")
        cell.font = Font(bold=True, size=10)
        cell.number_format = MONEY
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell = ws.cell(row=TS8, column=8, value="100%")
        cell.font = Font(bold=True, size=10)
        cell = ws.cell(row=TS9, column=7, value=f"=I{TR}/G{TR}")
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # ════════════════════════════════════════════════════════════════════
    # COMPROBANTES + DEVOLUCIONES (lado a lado)
    # ════════════════════════════════════════════════════════════════════
    # Row CS: headers de sección
    sec_hdr(ws, CS, 1, 3, "COMPROBANTES", level=1)
    cell = ws.cell(row=CS, column=4, value="TOTAL")
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE_MD)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    sec_hdr(ws, CS, 6, 7, "DEVOLUCIONES", level=1)
    ws.merge_cells(f"H{CS}:I{CS}")
    cell = ws.cell(row=CS, column=8, value="TOTAL")
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE_MD)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row CT: totales
    cell = ws.cell(row=CT, column=4, value=f"=SUM(D{CD1}:D{CDL})")
    cell.font = Font(bold=True, size=10)
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = PatternFill("solid", fgColor=GOLD)

    ws.merge_cells(f"H{CT}:I{CT}")
    cell = ws.cell(row=CT, column=8, value=f"=SUM(L{CD1}:L{CDL})")
    cell.font = Font(bold=True, size=10)
    cell.number_format = MONEY
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = PatternFill("solid", fgColor=GOLD)

    # Row CH: column headers
    for col, label in [(1,"FECHA"),(2,"CLIENTE"),(3,"CUIT"),(4,"MONTO")]:
        hdr_cell(ws, CH, col, label)
    for col, label in [(6,"FECHA"),(7,"CLIENTE"),(8,"CONS"),(9,"COD"),
                       (10,"UNID."),(11,"$ x UNID"),(12,"TOTAL")]:
        hdr_cell(ws, CH, col, label)
    ws.row_dimensions[CH].height = 22

    # Rows CD1:CDL — fills intercalados + fórmula TOTAL en col L
    FILL_ODD  = PatternFill("solid", fgColor=WHITE)
    FILL_EVEN = PatternFill("solid", fgColor=ROW_ALT)
    for row in range(CD1, CDL + 1):
        rfill = FILL_EVEN if row % 2 == 0 else FILL_ODD
        # COMPROBANTES: cols A-D
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = rfill
        # DEVOLUCIONES: cols F-L
        for col in range(6, 13):
            ws.cell(row=row, column=col).fill = rfill
        # Fórmula TOTAL devoluciones
        cell = ws.cell(row=row, column=12,
                       value=f'=IF(J{row}*K{row}<>0,J{row}*K{row},"")')
        cell.number_format = MONEY
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = rfill

    ws.freeze_panes = f"A{D1}"
    ws.row_dimensions[H].height = 28

    fname = build_filename("Mod Yaguar", semana)
    return stream_wb(wb, fname)


# ── Artículos catálogo export ──────────────────────────────────────────────────

def _export_articulos(mayorista: str):
    with get_db() as cur:
        cur.execute("""
            SELECT * FROM articulos_catalogo
            WHERE mayorista = %s
            ORDER BY descrip
        """, (mayorista,))
        rows = [dict(r) for r in cur.fetchall()]

    es_yaguar = mayorista == "yaguar"

    if es_yaguar:
        hdrs = [
            ("Código",        "cod_art",          14),
            ("Barcode",       "cod_bar",          16),
            ("Descripción",   "descrip",          45),
            ("Fabricante",    "fabricante",       22),
            ("UxB",           "uxb",               8),
            ("Precio c/IVA",  "precio_con_iva",   14),
            ("% IVA",         "porc_iva",         10),
            ("Costo",         "precio_costo",     14),
            ("% Dto.",        "descuento_default",10),
            ("Imp. Monto",    "impuestos_monto",  12),
            ("Imp. %",        "impuestos_porc",   10),
            ("Subcategoría",  "subcategoria",     28),
            ("Unidad med.",   "unidad_medida",    12),
            ("Stock",         "stock",            10),
            ("Estado",        "estado",            8),
            ("Observaciones", "observaciones",    30),
            ("Folder",        "folder",           16),
        ]
    else:
        hdrs = [
            ("Código",          "cod_art",          14),
            ("Barcode unidad",  "cod_bar",          16),
            ("Barcode bulto",   "cod_bar_bulto",    16),
            ("Descripción",     "descrip",          45),
            ("UxB",             "uxb",               8),
            ("Precio c/IVA",    "precio_con_iva",   14),
            ("Costo",           "precio_costo",     14),
            ("Precio mayorista","precio_mayorista", 16),
            ("Familia",         "familia",          20),
            ("Subcategoría",    "subcategoria",     28),
            ("Tipo unidad",     "tipo_unidad",      12),
            ("Estado cat.",     "tipo_estado",      12),
            ("Stock",           "stock",            10),
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Artículos"
    ws.row_dimensions[1].height = 28

    for ci, (label, _, width) in enumerate(hdrs, 1):
        hdr_cell(ws, 1, ci, label)
        ws.column_dimensions[get_column_letter(ci)].width = width

    FILL_ODD  = PatternFill("solid", fgColor=WHITE)
    FILL_EVEN = PatternFill("solid", fgColor=ROW_ALT)

    for ri, r in enumerate(rows, 2):
        row_fill = FILL_EVEN if ri % 2 == 0 else FILL_ODD
        for ci, (label, field, _) in enumerate(hdrs, 1):
            val = r.get(field)
            fmt, align = None, "left"
            if field in ("precio_con_iva", "precio_costo", "precio_mayorista",
                         "impuestos_monto"):
                fmt, align = MONEY, "right"
            elif field in ("porc_iva", "descuento_default", "impuestos_porc"):
                fmt, align = '0.00"%"', "center"
            elif field in ("uxb", "stock", "estado"):
                align = "center"
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=10, color=TEXT)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if fmt:
                cell.number_format = fmt

    if rows:
        last = len(rows) + 1
        make_table(ws, f"A1:{get_column_letter(len(hdrs))}{last}", "TablaArticulos")

    ws.freeze_panes = "A2"
    fname = build_filename("Articulos", mayorista.capitalize())
    return stream_wb(wb, fname)


# ── Rutas ──────────────────────────────────────────────────────────────────────

@router_yaguar.get("/picks")
def yaguar_export(semana: str = Query(...)):
    return _export_picks(semana, "yaguar")

@router_yaguar.get("/mod")
def yaguar_export_mod(semana: str = Query(...)):
    return _export_mod_yaguar(semana)

@router_yaguar.get("/clientes")
def yaguar_export_clientes():
    return _export_clientes("yaguar")

@router_diarco.get("/picks")
def diarco_export(semana: str = Query(...)):
    return _export_picks(semana, "diarco")

@router_diarco.get("/clientes")
def diarco_export_clientes():
    return _export_clientes("diarco")

@router_yaguar.get("/articulos")
def yaguar_export_articulos():
    return _export_articulos("yaguar")

@router_diarco.get("/articulos")
def diarco_export_articulos():
    return _export_articulos("diarco")
