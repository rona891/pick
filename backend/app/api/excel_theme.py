"""
Módulo compartido de estilos para todos los exports Excel.
Paleta: azul profesional (#1F3864 navy) con texto oscuro legible.
"""
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paleta de colores ─────────────────────────────────────────────────────────
NAVY     = "1F3864"   # header principal (azul navy oscuro)
WHITE    = "FFFFFF"   # fila impar (blanca pura)
ROW_ALT  = "C6D9F1"   # fila par (azul más marcado — mejor contraste)
GOLD     = "FFF9C4"   # fila TOTAL / resumen (amarillo suave)
BLUE_MD  = "2E75B6"   # encabezados de sección secundarios (azul medio)
BLUE_LT  = "BDD7EE"   # sub-encabezados / detalles (azul claro)
BORDER   = "C9C9C9"   # bordes (gris claro)
TEXT     = "1A1A1A"   # texto de datos (casi negro)
GRAY_BG  = "F2F2F2"   # gris muy suave (fuera de tabla)

# Status específicos del export Pick
S_OK     = "375623"   # texto estado completado (verde oscuro)
S_PT     = "833C00"   # texto estado parcial (naranja oscuro)
S_NO     = "595959"   # texto estado pendiente (gris)

# ── Formatos numéricos ────────────────────────────────────────────────────────
MONEY    = '"$"#,##0.00'
PCT      = '0.00%'
PCT_INT  = '0%'
INT_FMT  = '#,##0'

# ── Estilo de tabla ───────────────────────────────────────────────────────────
TABLE_STYLE = "TableStyleMedium2"   # azul profesional, stripes automáticos


# ── Helpers de borde ─────────────────────────────────────────────────────────

def thin_brd(color: str = BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def top_brd(color: str = BORDER) -> Border:
    """Solo borde superior (para separar fila TOTAL)."""
    t = Side(style="medium", color=NAVY)
    n = Side(style=None)
    return Border(top=t, left=n, right=n, bottom=n)


# ── Helpers de celda ──────────────────────────────────────────────────────────

def hdr_cell(ws, row: int, col: int, text: str) -> None:
    """Celda de encabezado: fondo navy, texto blanco, bold, centrado."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=True, size=10, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_brd(NAVY)


def data_cell(ws, row: int, col: int, value=None,
              align: str = "left", fmt: str = None,
              bold: bool = False, color: str = TEXT) -> None:
    """Celda de datos: sin fill (lo maneja la tabla), fuente oscura, formato opcional."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, size=10, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt


def tot_cell(ws, row: int, col: int, value=None,
             fmt: str = None, align: str = "right") -> None:
    """Celda de fila TOTAL: fondo dorado, texto oscuro bold."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, size=10, color=TEXT)
    cell.fill      = PatternFill("solid", fgColor=GOLD)
    cell.border    = top_brd()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt


def sec_hdr(ws, row: int, col1: int, col2: int, text: str,
            level: int = 1) -> None:
    """Encabezado de sección para el Mod (merge + color)."""
    bg = BLUE_MD if level == 1 else BLUE_LT
    tc = WHITE    if level == 1 else TEXT
    ws.merge_cells(
        f"{get_column_letter(col1)}{row}:{get_column_letter(col2)}{row}"
    )
    cell = ws.cell(row=row, column=col1, value=text)
    cell.font      = Font(bold=True, size=11 if level == 1 else 10, color=tc)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_brd(NAVY if level == 1 else BORDER)
    ws.row_dimensions[row].height = 20


def make_table(ws, ref: str, name: str,
               style: str = TABLE_STYLE,
               stripes: bool = True) -> Table:
    """Crea y registra una tabla Excel dinámica con stripes automáticos."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=stripes,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    return tbl


def set_col_widths(ws, widths: dict) -> None:
    """widths = {col_number_or_letter: width}"""
    for col, w in widths.items():
        letter = col if isinstance(col, str) else get_column_letter(col)
        ws.column_dimensions[letter].width = w


def build_filename(name: str, identifier: str) -> str:
    """Ej: build_filename('Pick', 'PICK 21-05-2026') → 'Pick PICK 21-05-2026.xlsx'"""
    return f"{name} {identifier}.xlsx"


def stream_wb(wb, filename: str):
    """Serializa el workbook y devuelve un StreamingResponse."""
    from fastapi.responses import StreamingResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
